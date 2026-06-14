#!/usr/bin/env python3
"""
Build a structured KF26 trepart data artifact.

The KF26 source package is a mix of Excel tables and PDF tables. The skovrejsning
profile from forudsætningsnotat #4 table 3.1 is not present as a normal Excel
table, so the audited table values are encoded here and expanded from the
compressed PDF layout (2033-2044 is one repeated annual column).

Input:
  data/kf26-hoering/kf26-datark-lulucf.xlsx

Output:
  data/kf26/trepart.json
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
SOURCE_DIR = REPO_ROOT / "data" / "kf26-hoering"
OUTPUT_PATH = REPO_ROOT / "data" / "kf26" / "trepart.json"
LULUCF_XLSX = SOURCE_DIR / "kf26-datark-lulucf.xlsx"

KF26_PAGE_URL = (
    "https://www.kefm.dk/klima/klimastatus-og-fremskrivning/"
    "klimastatus-og-fremskrivning-2026"
)
KF26_DEL1_PDF_URL = (
    "https://www.kefm.dk/Media/639141758684018930/"
    "Klimastatus-%20og%20fremskrivning%202026%20Del%201.pdf"
)
FORUDSAETNINGSNOTAT_PATH = "data/kf26-hoering/4-forudsaetningsnotat-landbrug-arealer-skov.pdf"
LULUCF_DATAARK_PATH = "data/kf26-hoering/kf26-datark-lulucf.xlsx"


def _series(default: float = 0.0) -> dict[int, float]:
    return {year: default for year in range(2025, 2048)}


def _set_values(target: dict[int, float], values: dict[int, float]) -> None:
    for year, value in values.items():
        target[year] = value


def build_skov_profile() -> list[dict[str, Any]]:
    """Expand KF26 forudsætningsnotat #4 table 3.1 to annual rows."""
    categories = {
        "stateOrdinaryHa": _series(),
        "stateUntouchedGtpHa": _series(),
        "privateSubsidyCapHa": _series(),
        "privateSubsidyGtpOrdinaryHa": _series(),
        "privateSubsidyGtpUntouchedHa": _series(),
        "klimaskovfondenHa": _series(),
    }

    _set_values(
        categories["stateOrdinaryHa"],
        {2025: 330, 2026: 330, 2027: 310, 2028: 250, 2029: 210, 2030: 210, 2031: 210},
    )
    _set_values(
        categories["stateUntouchedGtpHa"],
        {2027: 400, 2028: 500, 2029: 800, 2030: 1000, 2031: 1000, 2032: 1000},
    )
    for year in range(2033, 2048):
        categories["stateUntouchedGtpHa"][year] = 1020

    _set_values(categories["privateSubsidyCapHa"], {2025: 841, 2026: 1040, 2027: 109})

    _set_values(
        categories["privateSubsidyGtpOrdinaryHa"],
        {
            2026: 4289,
            2027: 6350,
            2028: 7988,
            2029: 6503,
            2030: 6972,
            2031: 7623,
            2032: 8062,
            2045: 8029,
            2046: 5095,
            2047: 2211,
        },
    )
    for year in range(2033, 2045):
        categories["privateSubsidyGtpOrdinaryHa"][year] = 8384

    _set_values(
        categories["privateSubsidyGtpUntouchedHa"],
        {
            2026: 2095,
            2027: 3103,
            2028: 3903,
            2029: 3177,
            2030: 3406,
            2031: 3724,
            2032: 3939,
            2045: 3926,
            2046: 2492,
            2047: 1083,
        },
    )
    for year in range(2033, 2045):
        categories["privateSubsidyGtpUntouchedHa"][year] = 4096

    _set_values(
        categories["klimaskovfondenHa"],
        {2025: 700, 2026: 660, 2027: 870, 2028: 820, 2029: 1000, 2030: 1200},
    )

    rows: list[dict[str, Any]] = []
    cumulative = 0.0
    for year in range(2025, 2048):
        row = {key: value_by_year[year] for key, value_by_year in categories.items()}
        total = sum(row.values())
        cumulative += total
        rows.append(
            {
                "year": year,
                **row,
                "totalNewInitiativesHa": round(total, 1),
                "cumulativeNewInitiativesHa": round(cumulative, 1),
            }
        )
    return rows


def extract_lulucf_kulstofrig_series() -> dict[str, Any]:
    """Read selected kulstofrig landbrugsjord area values from KF26 LULUCF dataark."""
    if not LULUCF_XLSX.exists():
        return {"error": f"Missing {LULUCF_DATAARK_PATH}"}

    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl is required to parse KF26 LULUCF dataark"}

    wb = openpyxl.load_workbook(LULUCF_XLSX, read_only=True, data_only=True)
    try:
        ws = wb["LULUCF_arealer"]
        header = next(ws.iter_rows(min_row=8, max_row=8, values_only=True))
        years = [int(v) for v in header[1:] if isinstance(v, int)]
        series: dict[int, float] | None = None
        for row in ws.iter_rows(min_row=9, values_only=True):
            label = row[0]
            if label == "I alt, > 6 % OC":
                values = row[1 : 1 + len(years)]
                series = {year: round(float(value), 1) for year, value in zip(years, values) if value is not None}
                break
    finally:
        wb.close()

    if not series:
        return {"error": "Could not find 'I alt, > 6 % OC' in LULUCF_arealer"}

    selected_years = [1990, 2024, 2025, 2030, 2033, 2035, 2050]
    return {
        "unit": "1000_ha",
        "label": "I alt, > 6 % OC",
        "selectedYears": {str(year): series[year] for year in selected_years if year in series},
        "sourceFile": LULUCF_DATAARK_PATH,
        "sourceSheet": "LULUCF_arealer",
    }


def main() -> None:
    skov_profile = build_skov_profile()
    total_skov = round(sum(row["totalNewInitiativesHa"] for row in skov_profile), 1)

    result = {
        "_meta": {
            "builtAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "sourcePageUrl": KF26_PAGE_URL,
            "inputFiles": [
                FORUDSAETNINGSNOTAT_PATH,
                LULUCF_DATAARK_PATH,
                "data/kf26-hoering/kapitel-20-landbrugsarealer.pdf",
                "data/kf26-hoering/kapitel-21-skov-traeprodukter.pdf",
            ],
            "notes": [
                "KF26 is a hearing version published 2026-05-12; final version expected 2026-09.",
                "Skov profile is expanded from forudsætningsnotat #4 table 3.1; the PDF groups 2033-2044 as one repeated annual value.",
            ],
        },
        "publishedAt": "2026-05-12",
        "version": "høring",
        "status": "I offentlig høring frem til 2026-06-12; endelig version forventes 2026-09",
        "sourceUrl": KF26_PAGE_URL,
        "targetsAndHorizons": {
            "politicalExtractionDeadline": "2030-12-31",
            "kf26ExtractionProjectAreaHorizon": "2032",
            "kf26ExtractionCarbonRichHorizon": "2033",
            "politicalAfforestationDeadline": "2045-12-31",
            "kf26AfforestationRealizationHorizon": "2047",
            "extractionProjectAreaTargetHa": 140000,
            "extractionAgriculturalAreaApproxHa": 120000,
            "extractionCarbonRichAgriculturalSoilTargetHa": 70000,
            "afforestationPoliticalTargetHa": 250000,
            "afforestationKf26ImkHa": 273000,
            "untouchedForestTargetWithinTrepartHa": 100000,
        },
        "lavbundStatusDec2025": {
            "underUdtagningHa": 71100,
            "forundersoegelsestilsagnHa": 52400,
            "vkpRunde3ForundersoegelseHa": 11800,
            "vkpRunde3EtableringHa": 2900,
            "definitionNote": "KF26-tallene dækker kulstofrige landbrugsarealer inkl. randarealer. MARS-totaler i trackeren dækker bredere projekttyper/faser og er derfor ikke direkte sammenlignelige.",
            "source": {
                "file": FORUDSAETNINGSNOTAT_PATH,
                "section": "Boks 2.3 / afsnit 2.4 Usikkerhed",
            },
        },
        "assumptions": {
            "lavbundNPlusYears": 5,
            "lavbundDropoutRateRange": [0.15, 0.35],
            "stateAfforestationRealization": "2 år efter tilsagn",
            "privateAfforestationRealization": "35% efter 1 år, 35% efter 2 år, 30% efter 3 år; 2025-tilsagn først fra 2026",
            "forestMineralSoilCarbonBindingYearsKf26": 100,
            "forestMineralSoilCarbonBindingYearsKf25": 30,
            "forestNetEffectIncrease2030MtCo2eVsKf25": 0.2,
            "forestNetEffectIncrease2035MtCo2eVsKf25": 0.2,
        },
        "skovProfilPerYear": skov_profile,
        "skovProfilSummary": {
            "sumNewInitiativesHa": total_skov,
            "roundedKf26ImkHa": 273000,
            "politicalTargetHa": 250000,
            "source": {
                "file": FORUDSAETNINGSNOTAT_PATH,
                "table": "Tabel 3.1",
            },
            "roundingNote": "Tabel 3.1 er afrundet til hele hektar; summen af de udfoldede år er derfor ca. 273.000 ha.",
        },
        "lulucfArealer": {
            "kulstofrigLandbrugsjord": extract_lulucf_kulstofrig_series(),
        },
        "landbrug2030Goal": {
            "sectorReductionPctKf26": 52,
            "lowerTargetPct": 55,
            "upperTargetPct": 65,
            "gapToLowerTargetMtCo2e": 0.7,
            "gapToUpperTargetMtCo2e": 2.8,
            "source": {
                "url": KF26_DEL1_PDF_URL,
                "section": "Tabel 1.1 / landbrugsmål i 2030",
            },
        },
        "co2Headline": {
            "target2030MarginMtCo2e": 0.4,
            "kf25Target2030MarginMtCo2e": 1.5,
            "source": {
                "url": KF26_DEL1_PDF_URL,
                "section": "Kapitel 1 / tabel 1.1",
            },
        },
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"✓ Wrote {OUTPUT_PATH.relative_to(REPO_ROOT)}")
    print(f"  Skov profile: {skov_profile[0]['year']}–{skov_profile[-1]['year']}, {total_skov:,.0f} ha")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"✗ {exc}", file=sys.stderr)
        sys.exit(1)
