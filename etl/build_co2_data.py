#!/usr/bin/env python3
"""
ETL: Build CO₂ emissions dashboard data from KF25 CRF tables.

Reads the KF25 CRF-tabeller Excel file (emissions by CRF sector and year)
and produces a compact JSON file with:

1. National emissions trajectory 1990–2050 by sector (energy, industry,
   agriculture, LULUCF, waste) — in million tonnes CO₂-equivalents
2. Agriculture sub-breakdown (enteric fermentation, manure, soils) —
   directly relevant to Trepart's CO₂e livestock tax
3. LULUCF sub-breakdown (forest, cropland, grassland, wetlands) —
   relevant to afforestation & land extraction pillars
4. 2030 target (70% reduction from 1990 excl. LULUCF)
5. Key milestones (last historic year, projected 2025/2030/2035)

Prerequisite: Run fetch_kf25.py first to download the source data.
Input:  data/kf25/kf25-crf-tabeller.xlsx
Output: public/data/co2-emissions.json
"""

import json
import sys
import time
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
DATA_DIR = REPO_ROOT / "data" / "kf25"
OUTPUT_PATH = REPO_ROOT / "public" / "data" / "co2-emissions.json"

CRF_FILE = DATA_DIR / "kf25-crf-tabeller.xlsx"

# CRF sector groupings
ENERGY_PREFIX = "1"
INDUSTRY_PREFIX = "2"
AGRICULTURE_PREFIX = "3"
LULUCF_PREFIX = "4"
WASTE_PREFIX = "5"
CORRECTION_CODES = {"ccs", "cor", "cor_5B2", "cor_5D1a"}

# Agriculture sub-categories for Trepart CO₂e tax tracking
AG_BREAKDOWN = {
    "entericFermentation": "3A",  # Cattle/livestock methane
    "manureManagement": "3B",     # Manure storage & handling
    "agriculturalSoils": "3D",    # N₂O from fertilizers/soils
}

# LULUCF sub-categories for afforestation & extraction tracking
LULUCF_BREAKDOWN = {
    "forestLand": "4A",   # Carbon sink from forests (negative = sequestration)
    "cropland": "4B",     # Emissions from cultivated organic soils
    "grassland": "4C",    # Grassland management
    "wetlands": "4D",     # Peatland/wetland emissions
}


def sum_sector(rows: dict, prefix: str, n_years: int) -> list[float]:
    """Sum all CRF rows starting with a given prefix."""
    total = [0.0] * n_years
    for code, data in rows.items():
        if code.startswith(prefix):
            for i in range(n_years):
                total[i] += data["values"][i]
    return [round(v, 3) for v in total]


def main():
    t0 = time.monotonic()
    print("CO₂ ETL — Building emissions data from KF25 CRF tables")

    if not CRF_FILE.exists():
        print(f"  ✗ CRF file not found: {CRF_FILE}")
        print("  Run: python3 etl/fetch_kf25.py")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("  ✗ openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    # --- Read CRF CO₂-equivalent sheet ---
    print(f"  Reading {CRF_FILE.name}...")
    wb = openpyxl.load_workbook(CRF_FILE, read_only=True, data_only=True)
    ws = wb["CO2-ækv"]

    # Extract years from header row (row 4)
    header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    years = [int(v) for v in header[2:] if v is not None]
    n_years = len(years)
    print(f"  Year range: {years[0]}–{years[-1]} ({n_years} years)")

    # Extract all data rows
    rows = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        code = row[0]
        if code:
            values = row[2 : 2 + n_years]
            rows[str(code)] = {
                "desc": str(row[1] or ""),
                "values": [float(v) if v is not None else 0.0 for v in values],
            }
    wb.close()
    print(f"  Parsed {len(rows)} CRF categories")

    # --- Compute sector totals ---
    energy = sum_sector(rows, ENERGY_PREFIX, n_years)
    industry = sum_sector(rows, INDUSTRY_PREFIX, n_years)
    agriculture = sum_sector(rows, AGRICULTURE_PREFIX, n_years)
    lulucf = sum_sector(rows, LULUCF_PREFIX, n_years)
    waste = sum_sector(rows, WASTE_PREFIX, n_years)

    # Correction terms (CCS, methodology corrections)
    corrections = [0.0] * n_years
    for code in CORRECTION_CODES:
        if code in rows:
            for i in range(n_years):
                corrections[i] += rows[code]["values"][i]

    # National total excl. LULUCF (basis for 70% target)
    total_excl = [
        round(energy[i] + industry[i] + agriculture[i] + waste[i] + corrections[i], 3)
        for i in range(n_years)
    ]
    # National total incl. LULUCF
    total_incl = [round(total_excl[i] + lulucf[i], 3) for i in range(n_years)]

    # --- Sub-breakdowns ---
    ag_breakdown = {}
    for key, code in AG_BREAKDOWN.items():
        if code in rows:
            ag_breakdown[key] = [round(rows[code]["values"][i], 3) for i in range(n_years)]

    lulucf_breakdown = {}
    for key, code in LULUCF_BREAKDOWN.items():
        if code in rows:
            lulucf_breakdown[key] = [round(rows[code]["values"][i], 3) for i in range(n_years)]

    # --- Determine last historic year (first year where projection flag appears) ---
    # In KF25 data, 2023 is the last year with actual inventory data
    last_historic = 2023

    # --- Compute milestones ---
    baseline_1990 = total_excl[0]
    target_2030 = round(baseline_1990 * 0.30, 2)  # 70% reduction = keep 30%

    def year_idx(y: int) -> int:
        return years.index(y)

    def reduction_pct(year: int) -> float:
        idx = year_idx(year)
        return round((1 - total_excl[idx] / baseline_1990) * 100, 1)

    milestones = {
        "lastHistoricYear": last_historic,
        "reduction2023Pct": reduction_pct(2023),
        "reduction2025Pct": reduction_pct(2025),
        "reduction2030Pct": reduction_pct(2030),
        "reduction2035Pct": reduction_pct(2035),
        "totalExcl2023": total_excl[year_idx(2023)],
        "totalExcl2030": total_excl[year_idx(2030)],
        "agriculture2023": agriculture[year_idx(2023)],
        "agriculture2030": agriculture[year_idx(2030)],
        "lulucf2023": lulucf[year_idx(2023)],
        "lulucf2030": lulucf[year_idx(2030)],
    }

    # --- Build output ---
    result = {
        "source": "KF25 - Klimastatus og -fremskrivning 2025 (KEFM)",
        "sourceUrl": "https://www.kefm.dk/klima/klimastatus-og-fremskrivning/klimastatus-og-fremskrivning-2025",
        "unit": "mio_ton_co2e",
        "years": years,
        "sectors": {
            "energy": energy,
            "industry": industry,
            "agriculture": agriculture,
            "lulucf": lulucf,
            "waste": waste,
        },
        "totals": {
            "exclLulucf": total_excl,
            "inclLulucf": total_incl,
        },
        "targets": {
            "baseline1990ExclLulucf": round(baseline_1990, 2),
            "target2030ExclLulucf": target_2030,
            "reductionPct": 70,
        },
        "agricultureBreakdown": ag_breakdown,
        "lulucfBreakdown": lulucf_breakdown,
        "milestones": milestones,
    }

    # --- Write output ---
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(result, f, separators=(",", ":"))

    size_kb = OUTPUT_PATH.stat().st_size / 1024
    elapsed = time.monotonic() - t0

    print(f"\n  ✓ Wrote {OUTPUT_PATH.name} ({size_kb:.0f} KB)")
    print(f"\n  === Key Milestones ===")
    print(f"  1990 baseline (excl. LULUCF): {baseline_1990:.2f} mio ton CO₂e")
    print(f"  2023 actual:  {milestones['totalExcl2023']:.2f} mio ton ({milestones['reduction2023Pct']}% reduction)")
    print(f"  2030 target:  {target_2030:.2f} mio ton (70% reduction)")
    print(f"  2030 proj:    {milestones['totalExcl2030']:.2f} mio ton ({milestones['reduction2030Pct']}% reduction)")
    print(f"  Gap to target: {milestones['totalExcl2030'] - target_2030:.2f} mio ton")
    print(f"\n  Agriculture: {milestones['agriculture2023']:.2f} → {milestones['agriculture2030']:.2f} mio ton")
    print(f"  LULUCF:      {milestones['lulucf2023']:.2f} → {milestones['lulucf2030']:.2f} mio ton")
    print(f"\n  Duration: {elapsed:.1f}s")


if __name__ == "__main__":
    main()
